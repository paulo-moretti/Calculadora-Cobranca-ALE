from openpyxl import load_workbook
import os
from datetime import datetime
import json
import xlwings as xw
import time

print("Calculando Juros...")

# Caminhos
script_dir = os.path.dirname(__file__)
Config = os.path.join(script_dir, 'Config')
Saida = os.path.join(script_dir, 'Saida')
juros_path = os.path.join(Config, 'juros_ale.xlsm')

# Carregar SELIC
selic_path = os.path.join(Config, 'selic.json')
with open(selic_path, 'r', encoding='utf-8') as f:
    selic_data = json.load(f)

mes_atual = datetime.now().strftime("%m/%Y")
selic_percentual = selic_data.get(mes_atual)
if selic_percentual is None:
    selic_percentual = float(input(f"Informe a SELIC para {mes_atual} (%): ").replace(',', '.'))
    selic_data[mes_atual] = selic_percentual
    with open(selic_path, 'w', encoding='utf-8') as f:
        json.dump(selic_data, f, indent=4, ensure_ascii=False)

# Planilhas temporárias
planilhas_temp = [
    f for f in os.listdir(Config)
    if f.endswith('_temp.xlsx') and not f.startswith('~$')
]
print(f"Planilhas temporárias encontradas: {planilhas_temp}")

# Processar cada planilha temporária
for planilha_temp in planilhas_temp:
    temp_path = os.path.join(Config, planilha_temp)

    # Abre com openpyxl para escrever no final
    wb = load_workbook(temp_path)
    ws = wb.active

    # Ler valores de J35:J45 da planilha temporária com xlwings
    app = xw.App(visible=False)
    app.calculation = 'automatic'  # força modo de cálculo automático

    wb_xlw = app.books.open(temp_path)
    ws_xlw = wb_xlw.sheets[0]
    valores_real = []

    for idx in range(11):
        valor_real = ws_xlw.range(f'J{35 + idx}').value
        print(f"Valor real de J{35 + idx}: {valor_real}")
        valores_real.append(valor_real if valor_real is not None else 0)

    wb_xlw.close()

    # Abrir planilha de juros, escrever em D81:D91, forçar cálculo e ler J81:J91
    try:
        wb_xlw_juros = app.books.open(juros_path)
        ws_xlw_juros = wb_xlw_juros.sheets[1]

        for idx in range(11):
            cell = ws_xlw_juros.range(f'D{81 + idx}')
            cell.value = valores_real[idx]
            print(f"Valor escrito em D{81 + idx}: {cell.value}")

        wb_xlw_juros.app.calculate()
        wb_xlw_juros.save()

        for idx in range(11):
            valor_juros = ws_xlw_juros.range(f'J{81 + idx}').value
            ws[f'N{35 + idx}'].value = valor_juros if valor_juros is not None else 0

        # Zerar o conteúdo da planilha de juros após o cálculo
        for idx in range(11):
            if valores_real[idx] != 0:  # Verifica se foi preenchido um valor
                cell = ws_xlw_juros.range(f'D{81 + idx}')
                cell.value = 0  # Zera o valor da célula

    finally:
        wb_xlw_juros.close()
        app.quit()

    # Atualizar SELIC
    ws['N48'].value = selic_percentual / 100
    ws['N48'].number_format = '0.00%'

    # Salvar planilha final
    nome_final = planilha_temp.replace('_temp.xlsx', '.xlsx')
    caminho_final = os.path.join(Saida, nome_final)
    wb.save(caminho_final)
    print(f"[✓] Planilha finalizada e salva: {caminho_final}")
