import os
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook

script_dir = os.path.dirname(__file__)
Config = os.path.join(script_dir, 'Config')
Saida = os.path.join(script_dir, 'Saida')

arquivos_json = [
    f for f in os.listdir(Config)
    if f.lower().endswith('.json') and f.lower() != 'selic.json'
]
if not arquivos_json:
    print("Nenhum arquivo JSON encontrado na pasta Config.")
    exit()

arquivos_xlsx = [f for f in os.listdir(Config) if f.lower().endswith('.xlsx')]
if not arquivos_xlsx:
    print("Nenhum arquivo XLSX encontrado na pasta Config.")
    exit()

planilha_modelo_path = os.path.join(Config, arquivos_xlsx[0])

def buscar_dados_json(valores_gratificacoes, data_excel):
    data_str = data_excel.strftime('01/%m/%Y')
    for registro in valores_gratificacoes:
        if registro.get('Data da Competência') == data_str:
            return registro
    return None

for json_file in arquivos_json:
    json_path = os.path.join(Config, json_file)

    with open(json_path, 'r', encoding='utf-8') as f:
        dados = json.load(f)

    valores_gratificacoes = dados['Valores das Gratificações']

    ale_50_str = dados.get('Valores Para fins de cálculo', {}).get('ALE 50%')
    ale_50_valor = float(ale_50_str.replace(',', '.')) if ale_50_str else 0.0

    nome = dados.get('Informações Pessoais', {}).get('Nome')
    cpf = dados.get('Informações Pessoais', {}).get('CPF')
    selic = dados.get('cell_t36')

    if not nome:
        print(f"Nome não encontrado no JSON: {json_file}")
        continue

    wb = load_workbook(planilha_modelo_path)
    ws = wb.active

    if nome:
        ws['E14'] = nome
    if cpf:
        ws['E15'] = cpf

    # Preencher coluna D com ALE 50%
    if ale_50_valor:
        for row in range(35, 46):
            cell = ws[f'D{row}']
            cell.value = round(ale_50_valor, 2)
            cell.number_format = '0.00'

    # Preencher coluna C com tempo de serviço conforme a data
    for row in range(35, 46):
        data_cell = ws[f'B{row}'].value
        if not data_cell:
            continue

        if isinstance(data_cell, datetime):
            data_excel = data_cell
        else:
            try:
                data_excel = datetime.strptime(str(data_cell), "%d/%m/%Y")
            except ValueError:
                print(f"Data inválida na célula C{row}: {data_cell}")
                continue

        registro_json = buscar_dados_json(valores_gratificacoes, data_excel)
        if registro_json:
            valor_c = registro_json.get('tempo_de_servico')
            if valor_c is not None:
                cell_c = ws[f'C{row}']
                valor_c_str = str(valor_c).replace(',', '.')
                valor_c_num = float(valor_c_str)
                cell_c.value = int(valor_c_num)
                cell_c.number_format = '0'

    # Preencher SELIC se existir
    if selic is not None:
        try:
            ws['N48'].value = selic / 100
            ws['N48'].number_format = '0.00%'
        except Exception as e:
            print(f"Erro ao preencher SELIC: {e}")

    if not os.path.exists(Saida):
        os.makedirs(Saida)

    nome_sanitizado = nome.replace(' ', '_')
    nome_arquivo_saida = os.path.join(Saida, f'{nome_sanitizado}_planilha.xlsx')
    def str_para_data(s):
        return datetime.strptime(s, "%d/%m/%Y")

    reajustes_json = dados.get("Valores Para fins de cálculo", {}).get("Reajustes Percentuais", [])

    faixas = []
    for i in range(len(reajustes_json)):
        inicio = str_para_data(reajustes_json[i]["de"])
        fim = str_para_data(reajustes_json[i + 1]["de"]) if i + 1 < len(reajustes_json) else datetime.max
        percentual = reajustes_json[i]["reajuste_percentual"]
        faixas.append((inicio, fim, percentual))
            
    for row in range(35, 46):
        celula_data = ws[f'B{row}'].value
        if not celula_data:
            continue

        try:
            data = celula_data if isinstance(celula_data, datetime) else datetime.strptime(str(celula_data), "%d/%m/%Y")
        except ValueError:
            print(f"[!] Data inválida na célula B{row}: {celula_data}")
            continue

        percentual_aplicado = ""
        for inicio, fim, perc in faixas:
            if inicio <= data < fim:
                percentual_aplicado = perc
                break

        if percentual_aplicado:
            perc_float = float(percentual_aplicado.replace('%', '').replace(',', '.')) / 100
            ws[f'E{row}'].value = perc_float
            ws[f'E{row}'].number_format = '0.00%'

    nome_sanitizado = nome.replace(' ', '_')
    planilha_temporaria = os.path.join(Config, f'{nome_sanitizado}_temp.xlsx')
    wb.save(planilha_temporaria)
    print(f"[✓] Planilha temporária preenchida: {planilha_temporaria}")
